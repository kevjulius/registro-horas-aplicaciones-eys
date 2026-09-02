from __future__ import annotations

import argparse
from pathlib import Path

import openpyxl


REQUIRED_HEADERS = [
    "Aplicativo",
    "Empresa",
    "Cant. horas",
    "Equipo",
    "Material CF",
    "Glosa PL",
]


def sql_quote(value: object) -> str:
    return "'" + str(value).replace("'", "''") + "'"


def numeric_sql(value: float) -> str:
    text = f"{value:.6f}".rstrip("0").rstrip(".")
    return text or "0"


def load_rows(path: Path, year: int) -> list[tuple[int, str, str, str, float, str, str]]:
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    headers = [str(cell).strip() if cell is not None else "" for cell in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))]
    missing = [header for header in REQUIRED_HEADERS if header not in headers]
    if missing:
        raise ValueError(f"Faltan columnas requeridas: {', '.join(missing)}")

    rows: list[tuple[int, str, str, str, float, str, str]] = []
    for raw in sheet.iter_rows(min_row=2, values_only=True):
        if not any(value is not None and str(value).strip() for value in raw):
            continue
        item = dict(zip(headers, raw))
        rows.append(
            (
                year,
                str(item["Equipo"]).strip(),
                str(item["Aplicativo"]).strip(),
                str(item["Empresa"]).strip(),
                float(item["Cant. horas"]),
                str(item["Material CF"]).strip(),
                str(item["Glosa PL"]).strip(),
            )
        )
    return rows


def build_sql(rows: list[tuple[int, str, str, str, float, str, str]], year: int) -> str:
    lines = [
        "-- Agrega presupuesto mensual fijo por sistema/sociedad para reporte Presupuesto vs Consumo.",
        "-- Ejecutar en Supabase SQL Editor.",
        "",
        "create table if not exists public.application_budgets (",
        "  id bigint generated always as identity primary key,",
        "  anio integer not null check (anio >= 2000),",
        "  equipo text not null check (equipo in ('Aplicaciones', 'BI')),",
        "  sistema text not null,",
        "  sociedad text not null,",
        "  horas_presupuestadas_mes numeric(12,3) not null check (horas_presupuestadas_mes >= 0),",
        "  material_cf text not null default '',",
        "  glosa_pl text not null default '',",
        "  active boolean not null default true,",
        "  created_at timestamptz not null default now(),",
        "  updated_at timestamptz not null default now()",
        ");",
        "",
        "create index if not exists application_budgets_lookup_idx",
        "  on public.application_budgets (anio, equipo, sistema, sociedad)",
        "  where active = true;",
        "",
        "alter table public.application_budgets enable row level security;",
        "",
        'drop policy if exists "application budgets read admin" on public.application_budgets;',
        'create policy "application budgets read admin" on public.application_budgets',
        "for select using (",
        "  public.current_profile_role() = 'administracion'",
        "  or (public.current_profile_role() = 'adminbi' and equipo = 'BI')",
        ");",
        "",
        f"-- Reemplaza la carga {year} para que el script sea reejecutable.",
        f"delete from public.application_budgets where anio = {year};",
        "",
        "insert into public.application_budgets",
        "  (anio, equipo, sistema, sociedad, horas_presupuestadas_mes, material_cf, glosa_pl)",
        "values",
    ]

    values = []
    for row in rows:
        values.append(
            "  ("
            + ", ".join(
                [
                    str(row[0]),
                    sql_quote(row[1]),
                    sql_quote(row[2]),
                    sql_quote(row[3]),
                    numeric_sql(row[4]),
                    sql_quote(row[5]),
                    sql_quote(row[6]),
                ]
            )
            + ")"
        )
    lines.append(",\n".join(values) + ";")
    return "\n".join(lines) + "\n"


def main() -> None:
    parser = argparse.ArgumentParser(description="Genera SQL de presupuestos desde Excel.")
    parser.add_argument("excel", type=Path)
    parser.add_argument("--year", type=int, default=2026)
    parser.add_argument("--output", type=Path, default=Path("supabase/add_application_budgets.sql"))
    args = parser.parse_args()

    rows = load_rows(args.excel, args.year)
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(build_sql(rows, args.year), encoding="utf-8")
    print(f"Generado {args.output} con {len(rows)} filas.")


if __name__ == "__main__":
    main()
